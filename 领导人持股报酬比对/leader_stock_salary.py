"""
领导人持股报酬比对系统
"""
import concurrent.futures
import json
import os
import sys
import threading
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Union
import pandas as pd

from config import validate_config
from database_manager import db_manager
from ai_service_enhanced import enhanced_ai_service
from logger_config import setup_logging, get_logger, get_session_id, get_file_only_logger

setup_logging()
logger = get_logger(__name__)
file_only_logger = get_file_only_logger(__name__)


def get_base_path():
    """获取程序基础路径，兼容开发和打包环境"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

PROMPT_FILE = "prompt_ldrcg.md"

SALARY_FIELDS_TO_COMPARE = [
    "实际所得报酬总额起始",
    "实际所得报酬总额截止",
    "实际所得报酬总额及人数"
]

HOLD_FIELDS_TO_COMPARE = [
    "职位描述",
    "期初持股数",
    "期末持股数",
    "间接持股数",
    "从公司获得年度报酬总额",
    "补贴津贴",
    "变动原因说明",
    "是否在股东或关联单位领取报酬津贴",
    "在任与否"
]

SQL_QUERY_HOLD = '''
SELECT B.GPDM,CONVERT(DATE,A.XXFBSJ) XXFBRQ,A.LDXM 领导人姓名,A.ZWMS 职位描述,
    A.QCCGS 期初持股数,A.CGS 期末持股数,A.JJCG 间接持股数,
    A.NDBC 从公司获得年度报酬总额,A.BTJT 补贴津贴,A.BDYYSM 变动原因说明,
    CASE WHEN A.SFZGDHGLDWLQBCJT=1 THEN '是' WHEN A.SFZGDHGLDWLQBCJT=2 THEN '否' ELSE '' END 是否在股东或关联单位领取报酬津贴,
    CASE WHEN A.ZRYF=1 THEN '是' WHEN A.ZRYF=2 THEN '否' ELSE '' END 在任与否
FROM [10.101.0.212].JYPRIME.dbo.usrZYLDRCG A
  LEFT JOIN [10.101.0.212].JYPRIME.dbo.usrZQZB B ON A.INBBM=B.INBBM AND B.ZQSC IN (83,90,18) AND B.ZQLB IN (1,2,41)
WHERE A.XXLYBM = 110101 AND B.GPDM = ? AND A.XXFBSJ = ?
ORDER BY XH
'''

SQL_QUERY_SALARY = '''
SELECT B.GPDM,CONVERT(DATE,A.XXFBRQ) XXFBRQ,
    A.BCQJQS 实际所得报酬总额起始,
    A.BCQJJZ 实际所得报酬总额截止,
    A.SJRS 实际所得报酬总额涉及人数
FROM [10.101.0.212].JYPRIME.dbo.usrGSJYLDRBC A 
    LEFT JOIN [10.101.0.212].JYPRIME.dbo.usrZQZB B ON A.IGSDM=B.IGSDM AND B.ZQSC IN (83,90,18) AND B.ZQLB IN (1,2,41)
WHERE A.XXLY='年度报告' AND A.BCXMLB = 6 AND B.GPDM = ? AND A.XXFBRQ = ?
'''


class LeaderStockSalaryProcessor:
    """领导人持股报酬数据处理类"""

    def __init__(self):
        self.lock = threading.Lock()
        self.file_status = {}
        self.uploaded_file_ids = {}

        try:
            validate_config()
        except ValueError as e:
            print(f"配置验证失败: {e}")
            raise

    def _get_company_key(self, filename: str) -> Optional[str]:
        """
        从文件名提取公司唯一标识（股票代码-发布日期）
        
        Args:
            filename: 文件名（不含路径）
            
        Returns:
            公司标识字符串，格式为"股票代码-发布日期"，失败返回None
        """
        try:
            base_name = os.path.splitext(filename)[0]
            parts = base_name.split('-')
            
            if len(parts) >= 4:
                stock_code = parts[0]
                publish_date = '-'.join(parts[1:4])
                
                try:
                    datetime.strptime(publish_date, '%Y-%m-%d')
                except ValueError:
                    logger.warning(f"日期格式错误: {publish_date}")
                    return None
                
                return f"{stock_code}-{publish_date}"
            else:
                logger.warning(f"文件名格式错误: {filename}")
                return None
                
        except Exception as e:
            logger.error(f"提取公司标识异常: {filename} - {e}")
            return None

    def process_all_files(self, base_dir: Path) -> List[Dict[str, Any]]:
        """处理所有PDF文件"""
        if not base_dir or not base_dir.exists():
            logger.error(f"目录不存在: {base_dir}")
            return []

        pdf_files = list(base_dir.glob("*.pdf"))
        if not pdf_files:
            logger.warning(f"未找到可处理的文件: {base_dir}")
            return []

        return self._pipeline_upload_and_process(pdf_files)

    def _pipeline_upload_and_process(self, pdf_files: List[Path]) -> List[Dict[str, Any]]:
        """
        流水线处理：上传和处理并行进行
        """
        all_results = []
        total_files = len(pdf_files)
        
        upload_workers = 2
        process_workers = 16
        
        upload_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=upload_workers,
            thread_name_prefix="Upload"
        )
        process_executor = concurrent.futures.ThreadPoolExecutor(
            max_workers=process_workers,
            thread_name_prefix="Process"
        )

        upload_futures = {}
        process_futures = {}
        
        upload_queue = list(pdf_files)

        upload_count = 0
        completed_count = 0
        file_ids = {}
        failed_uploads = set()

        print(f"任务流配置: {upload_workers}个上传线程, {process_workers}个处理线程")
        print(f"共 {total_files} 个文件待处理")

        try:
            while upload_queue or upload_futures or process_futures:
                while len(upload_futures) < upload_workers and upload_queue:
                    pdf_file = upload_queue.pop(0)
                    
                    future = upload_executor.submit(
                        self._upload_single_file_with_timeout,
                        pdf_file
                    )
                    upload_futures[future] = pdf_file
                    logger.debug(f"提交上传任务: {pdf_file.name}")

                if upload_futures:
                    completed_uploads = []
                    try:
                        for future in concurrent.futures.as_completed(upload_futures, timeout=0.1):
                            completed_uploads.append(future)
                    except concurrent.futures.TimeoutError:
                        pass
                    
                    for future in completed_uploads:
                        pdf_file = upload_futures.pop(future)

                        try:
                            file_id = future.result()

                            if file_id:
                                file_ids[pdf_file] = {
                                    "file_id": file_id,
                                    "filename": pdf_file.name,
                                    "file_path": pdf_file
                                }
                                upload_count += 1
                                print(f"↑ 上传成功({upload_count}/{total_files}): {pdf_file.name}")
                                
                                process_future = process_executor.submit(
                                    self._process_and_cleanup_file,
                                    pdf_file,
                                    file_ids[pdf_file]
                                )
                                process_futures[process_future] = pdf_file
                                logger.debug(f"提交处理任务: {pdf_file.name}")
                            else:
                                failed_uploads.add(pdf_file.name)
                                print(f"✗ 上传失败: {pdf_file.name}")

                        except Exception as e:
                            failed_uploads.add(pdf_file.name)
                            print(f"✗ 上传异常: {pdf_file.name} - {e}")

                if process_futures:
                    completed_processes = []
                    try:
                        for future in concurrent.futures.as_completed(process_futures, timeout=0.1):
                            completed_processes.append(future)
                    except concurrent.futures.TimeoutError:
                        pass
                    
                    for future in completed_processes:
                        pdf_file = process_futures.pop(future)

                        try:
                            result = future.result()

                            if result:
                                if result.get("status") == "no_data":
                                    status = "无数据"
                                    status_symbol = "○"
                                else:
                                    status = "成功"
                                    status_symbol = "✓"
                                all_results.append(result)
                            else:
                                status = "失败"
                                status_symbol = "✗"

                            completed_count += 1
                            print(f"{status_symbol} 处理{status}({completed_count}/{total_files}): {pdf_file.name}")

                        except Exception as e:
                            completed_count += 1
                            print(f"✗ 处理异常({completed_count}/{total_files}): {pdf_file.name} - {e}")

                time.sleep(0.05)

        except Exception as e:
            logger.error(f"流水线处理过程中发生异常: {e}", exc_info=True)
        
        finally:
            logger.info("开始关闭线程池...")

            for future in list(upload_futures.keys()):
                try:
                    future.result(timeout=30)
                except Exception as e:
                    logger.error(f"等待上传任务完成时出错: {e}")
                    future.cancel()

            for future in list(process_futures.keys()):
                try:
                    future.result(timeout=120)
                except Exception as e:
                    logger.error(f"等待处理任务完成时出错: {e}")
                    future.cancel()

            upload_executor.shutdown(wait=True)
            process_executor.shutdown(wait=True)

            logger.info("所有线程池已关闭")

        return all_results

    def _process_and_cleanup_file(self, pdf_file: Path, 
                                   file_info: Dict) -> Optional[Dict[str, Any]]:
        """
        处理单个文件并清理资源
        """
        process_start_time = time.time()
        logger.info(f"开始处理文件: {pdf_file.name}")
        result = None

        try:
            result = self._process_single_file(pdf_file, file_info)

            process_duration = time.time() - process_start_time
            if result:
                if result.get("status") == "no_data":
                    with self.lock:
                        self.file_status[pdf_file] = "no_data"
                    file_only_logger.info(f"文件无披露数据: {pdf_file.name} (耗时: {process_duration:.2f}秒)")
                else:
                    with self.lock:
                        self.file_status[pdf_file] = "completed"
                    file_only_logger.info(f"文件处理成功: {pdf_file.name} (耗时: {process_duration:.2f}秒)")
            else:
                logger.warning(f"文件处理失败: {pdf_file.name} (耗时: {process_duration:.2f}秒)")
                with self.lock:
                    self.file_status[pdf_file] = "failed"

            return result

        except Exception as e:
            process_duration = time.time() - process_start_time
            logger.error(f"处理文件异常: {pdf_file.name} (耗时: {process_duration:.2f}秒) - 错误: {str(e)}", exc_info=True)
            with self.lock:
                self.file_status[pdf_file] = "error"
            return None

        finally:
            try:
                self._cleanup_single_file(file_info["file_id"])
                file_only_logger.debug(f"已清理上传文件: {file_info['filename']}")
            except Exception as e:
                logger.error(f"清理上传文件失败: {file_info['filename']} - 错误: {str(e)}")

    def _preprocess_position(self, position: str) -> str:
        """
        预处理职位描述字段
        
        Args:
            position: 原始职位描述
            
        Returns:
            处理后的职位描述
        """
        import re
        value = position.replace(" ", "")
        for keyword in ["离任", "前任", "前", "原"]:
            value = value.replace(keyword, "")
        value = value.replace("（", "(").replace("）", ")")
        value = value.replace("兼", "、")
        value = value.replace("董秘", "董事会秘书")
        value = value.replace("高管", "高级管理人员")
        value = re.sub(r'副总(?!经理|裁|监)', '副总经理', value)
        return value

    def _is_all_english(self, text: str) -> bool:
        """
        判断字符串是否全为英文字符
        
        Args:
            text: 待判断的字符串
            
        Returns:
            是否全为英文
        """
        import re
        return bool(re.match(r'^[a-zA-Z\s]+$', text))

    def _preprocess_leader_name(self, name: str) -> str:
        """
        预处理领导人姓名
        
        Args:
            name: 原始姓名
            
        Returns:
            处理后的姓名
        """
        value = name.strip()
        value = value.replace("（", "(").replace("）", ")")
        if self._is_all_english(value):
            import re
            value = re.sub(r'\s+', ' ', value)
        else:
            value = value.replace(" ", "")
        return value

    def _get_compare_key(self, name: str) -> str:
        """
        获取用于比对的键（统一小写+去空格）
        
        Args:
            name: 姓名
            
        Returns:
            比对用的键
        """
        return self._preprocess_leader_name(name).lower()

    def _preprocess_ai_data(self, ai_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        预处理AI数据
        
        Args:
            ai_data: AI提取的数据
            
        Returns:
            预处理后的数据
        """
        result = {
            "Leader_Salary_data": None,
            "Leader_Hold_data": []
        }
        
        salary_data = ai_data.get("Leader_Salary_data", {})
        if salary_data:
            processed_salary = {}
            for field in SALARY_FIELDS_TO_COMPARE:
                value = str(salary_data.get(field, "")).strip()
                if value in ("不适用", "-", "/", "－", "N/A", "n/a", ""):
                    processed_salary[field] = ""
                else:
                    value = value.replace(" ", "").replace(",", "").replace("，", "")
                    processed_salary[field] = value
            result["Leader_Salary_data"] = processed_salary
        
        hold_data_list = ai_data.get("Leader_Hold_data", [])
        if hold_data_list:
            for record in hold_data_list:
                processed_record = {}
                original_ldxm = str(record.get("领导人姓名", "")).strip()
                processed_ldxm = self._preprocess_leader_name(original_ldxm)
                processed_record["领导人姓名"] = processed_ldxm
                processed_record["_original_name"] = original_ldxm
                
                for field in HOLD_FIELDS_TO_COMPARE:
                    value = str(record.get(field, "")).strip()
                    if value in ("不适用", "-", "/", "－", "N/A", "n/a", " "):
                        processed_record[field] = ""
                    else:
                        if field == "职位描述":
                            value = self._preprocess_position(value)
                        elif field == "变动原因说明":
                            if value in ("不适用", "-", "—", "/", "无", "不涉及", "无变化"):
                                value = ""
                            else:
                                value = value.replace(" ", "")
                        elif field in ["期初持股数", "期末持股数", "间接持股数"]:
                            value = value.replace(" ", "").replace(",", "").replace("，", "")
                            if "." in value:
                                try:
                                    num_value = float(value)
                                    value = str(int(num_value))
                                except ValueError:
                                    pass
                        elif field in ["从公司获得年度报酬总额", "补贴津贴"]:
                            value = value.replace(" ", "").replace(",", "").replace("，", "")
                        else:
                            value = value.replace(" ", "")
                        processed_record[field] = value
                
                if processed_ldxm:
                    result["Leader_Hold_data"].append(processed_record)
        
        return result

    def _process_single_file(self, pdf_file: Path, file_info: Dict) -> Optional[Dict[str, Any]]:
        """
        处理单个PDF文件，提取数据并进行比对
        """
        stock_code, publish_date = self._parse_filename(pdf_file.name)
        if not stock_code or not publish_date:
            logger.error(f"文件名解析失败: {pdf_file.name}")
            return None
        
        ai_data = None
        
        try:
            ai_data = self._extract_data(file_info["file_id"], pdf_file.name)
            
            if not ai_data:
                logger.error(f"AI数据提取失败: {pdf_file.name}")
                return None
            
            try:
                base_dir = get_base_path()
                log_dir = os.path.join(base_dir, "logs")
                if not os.path.exists(log_dir):
                    os.makedirs(log_dir)
                session_id = get_session_id()
                with open(os.path.join(log_dir, f"ai_extraction_data_{session_id}.log"), "a", encoding="utf-8") as f:
                    f.write(f"\n=== {pdf_file.name} ===\n")
                    f.write(f"{json.dumps(ai_data, ensure_ascii=False, indent=2)}\n")
            except Exception:
                pass
            
            ai_data = self._preprocess_ai_data(ai_data)
            
            if not ai_data["Leader_Salary_data"] and not ai_data["Leader_Hold_data"]:
                return {"status": "no_data", "pdf_file_name": pdf_file.name}

            sql_hold_data = None
            sql_salary_data = None
            try:
                sql_hold_data = self._query_database_hold(stock_code, publish_date)
                sql_salary_data = self._query_database_salary(stock_code, publish_date)
            except Exception as e:
                logger.warning(f"数据库查询失败: {pdf_file.name} - {str(e)}")

            comparison_result = self._compare_data(ai_data, sql_hold_data, sql_salary_data, stock_code, publish_date)

            result = {
                "stock_code": stock_code,
                "publish_date": publish_date,
                "pdf_file_path": str(pdf_file),
                "pdf_file_name": pdf_file.name,
                "ai_data": ai_data,
                "sql_hold_data": sql_hold_data,
                "sql_salary_data": sql_salary_data,
                "comparison_result": comparison_result
            }

            logger.info(f"处理成功: {pdf_file.name}")
            return result

        except Exception as e:
            logger.error(f"处理文件异常: {pdf_file.name} - 错误: {str(e)}")
            return None

    def _cleanup_single_file(self, file_id: str):
        """清理单个上传的文件"""
        try:
            enhanced_ai_service.delete_file(file_id)
        except Exception as e:
            pass

    def _upload_single_file_with_timeout(self, pdf_file: Path) -> Optional[str]:
        """
        上传单个PDF文件到AI平台
        
        参数:
            pdf_file: PDF文件路径
            
        返回:
            file_id: 上传成功后的文件ID，失败返回None
        """
        max_retries = 1
        
        for retry_count in range(max_retries + 1):
            try:
                file_id = enhanced_ai_service.upload_file(pdf_file)

                with self.lock:
                    self.uploaded_file_ids[pdf_file] = file_id
                
                if retry_count == 0:
                    file_size = pdf_file.stat().st_size / 1024 / 1024
                    file_only_logger.info(f"上传文件: {pdf_file.name} (大小: {file_size:.2f} MB, ID: {file_id})")

                return file_id

            except Exception as e:
                if retry_count == max_retries:
                    logger.error(f"文件上传失败: {pdf_file.name} - 错误: {str(e)}")
                    return None
                
                time.sleep(1)

    def _query_database_hold(self, stock_code: str, publish_date: str) -> Optional[List[Dict[str, Any]]]:
        """查询数据库获取领导人持股数据"""
        try:
            sql_data = db_manager.execute_query(SQL_QUERY_HOLD, (stock_code, publish_date))

            if not sql_data:
                file_only_logger.info(f"未找到持股数据: {stock_code} {publish_date}")
                return None

            results = []
            for record in sql_data:
                result = {}
                for key, value in record.items():
                    result[key] = value if value is not None else ""
                results.append(result)

            file_only_logger.info(f"找到持股数据: {stock_code} {publish_date} ({len(results)}条)")
            return results

        except Exception as e:
            logger.error(f"数据库查询持股失败: {stock_code} {publish_date} - 错误: {str(e)}")
            return None

    def _query_database_salary(self, stock_code: str, publish_date: str) -> Optional[Dict[str, Any]]:
        """查询数据库获取领导人报酬数据"""
        try:
            sql_data = db_manager.execute_query(SQL_QUERY_SALARY, (stock_code, publish_date))

            if not sql_data:
                file_only_logger.info(f"未找到报酬数据: {stock_code} {publish_date}")
                return None

            result = {}
            for key, value in sql_data[0].items():
                result[key] = value if value is not None else ""

            file_only_logger.info(f"找到报酬数据: {stock_code} {publish_date}")
            return result

        except Exception as e:
            logger.error(f"数据库查询报酬失败: {stock_code} {publish_date} - 错误: {str(e)}")
            return None

    def _parse_filename(self, filename: str) -> tuple:
        """解析文件名，提取股票代码、信息发布日期"""
        try:
            base_name = os.path.splitext(filename)[0]
            parts = base_name.split('-')

            if len(parts) >= 4:
                stock_code = parts[0]
                publish_date = '-'.join(parts[1: 4])

                try:
                    datetime.strptime(publish_date, '%Y-%m-%d')
                except ValueError:
                    print(f"日期格式错误: {publish_date}")
                    return None, None

                return stock_code, publish_date
            else:
                print(f"文件名格式错误: {filename}")
                return None, None

        except Exception as e:
            print(f"解析文件名异常: {e}")
            return None, None

    def _compare_values(self, value1: Any, value2: Any) -> bool:
        """比较两个值是否相等"""
        try:
            if not value1 and not value2:
                return True
            if not value1 or not value2:
                return False

            str_value1 = str(value1).strip()
            str_value2 = str(value2).strip()

            return str_value1 == str_value2
        except Exception:
            return False

    def _compare_data(self, ai_data: Dict[str, Any],
                      sql_hold_data: List[Dict[str, Any]],
                      sql_salary_data: Dict[str, Any],
                      stock_code: str, publish_date: str) -> List[Dict[str, Any]]:
        """
        比对领导人持股和报酬数据
        
        Args:
            ai_data: AI提取的数据
            sql_hold_data: SQL查询的持股数据列表
            sql_salary_data: SQL查询的报酬数据
            stock_code: 股票代码
            publish_date: 信息发布日期
            
        Returns:
            比对结果列表
        """
        results = []
        
        salary_comparison = self._compare_salary_data(
            ai_data.get("Leader_Salary_data"),
            sql_salary_data,
            stock_code,
            publish_date
        )
        if salary_comparison:
            results.append(salary_comparison)
        
        hold_comparisons = self._compare_hold_data(
            ai_data.get("Leader_Hold_data", []),
            sql_hold_data,
            stock_code,
            publish_date
        )
        results.extend(hold_comparisons)
        
        return results

    def _compare_salary_data(self, ai_salary_data: Optional[Dict[str, Any]],
                             sql_salary_data: Optional[Dict[str, Any]],
                             stock_code: str, publish_date: str) -> Optional[Dict[str, Any]]:
        """
        比对报酬数据
        
        Args:
            ai_salary_data: AI提取的报酬数据
            sql_salary_data: SQL查询的报酬数据
            stock_code: 股票代码
            publish_date: 信息发布日期
            
        Returns:
            比对结果
        """
        if not ai_salary_data and not sql_salary_data:
            return None
        
        if not sql_salary_data:
            return {
                "GPDM": stock_code,
                "XXFBRQ": publish_date,
                "领导人姓名": "报酬比对",
                "比对结果": "正式库无报酬数据"
            }
        
        if not ai_salary_data:
            return {
                "GPDM": stock_code,
                "XXFBRQ": publish_date,
                "领导人姓名": "报酬比对",
                "比对结果": "AI无报酬数据"
            }
        
        error_messages = []
        
        for field_name in SALARY_FIELDS_TO_COMPARE:
            ai_value = str(ai_salary_data.get(field_name, "")).strip()
            sql_value = str(sql_salary_data.get(field_name, "")).strip()
            
            ai_processed = self._preprocess_value(ai_value)
            sql_processed = self._preprocess_value(sql_value)
            
            if not self._compare_values(ai_processed, sql_processed):
                error_messages.append(f"{field_name}错误【正式库：{sql_value}，AI：{ai_value}】")
        
        comparison_result = "数据一致" if not error_messages else "\n".join(error_messages)
        
        return {
            "GPDM": stock_code,
            "XXFBRQ": publish_date,
            "领导人姓名": "报酬比对",
            "比对结果": comparison_result
        }

    def _compare_hold_data(self, ai_hold_data: List[Dict[str, Any]],
                           sql_hold_data: Optional[List[Dict[str, Any]]],
                           stock_code: str, publish_date: str) -> List[Dict[str, Any]]:
        """
        比对持股数据
        
        Args:
            ai_hold_data: AI提取的持股数据列表
            sql_hold_data: SQL查询的持股数据列表
            stock_code: 股票代码
            publish_date: 信息发布日期
            
        Returns:
            比对结果列表
        """
        results = []
        
        if not sql_hold_data:
            for ai_data in ai_hold_data:
                original_name = ai_data.get("_original_name", ai_data.get("领导人姓名", ""))
                results.append({
                    "GPDM": stock_code,
                    "XXFBRQ": publish_date,
                    "领导人姓名": original_name,
                    "比对结果": "正式库无对应记录"
                })
            return results
        
        sql_data_by_key = {}
        sql_original_names = {}
        for record in sql_hold_data:
            ldxm = str(record.get("领导人姓名", "")).strip()
            key = self._get_compare_key(ldxm)
            if key not in sql_data_by_key:
                sql_data_by_key[key] = record
                sql_original_names[key] = ldxm
        
        matched_sql_keys = set()
        
        for ai_data in ai_hold_data:
            ai_ldxm = str(ai_data.get("领导人姓名", "")).strip()
            ai_original_name = ai_data.get("_original_name", ai_ldxm)
            ai_key = self._get_compare_key(ai_ldxm)
            
            if ai_key not in sql_data_by_key.keys():
                results.append({
                    "GPDM": stock_code,
                    "XXFBRQ": publish_date,
                    "领导人姓名": ai_original_name,
                    "比对结果": "正式库无对应主键的记录"
                })
            else:
                matched_sql_keys.add(ai_key)
                
                sql_record = sql_data_by_key[ai_key]
                comparison_result = self._compare_hold_fields(ai_data, sql_record)
                
                results.append({
                    "GPDM": stock_code,
                    "XXFBRQ": publish_date,
                    "领导人姓名": ai_original_name,
                    "比对结果": comparison_result
                })
        
        for sql_key, sql_record in sql_data_by_key.items():
            if sql_key not in matched_sql_keys:
                sql_original_name = sql_original_names.get(sql_key, "")
                results.append({
                    "GPDM": stock_code,
                    "XXFBRQ": publish_date,
                    "领导人姓名": sql_original_name,
                    "比对结果": "AI无对应记录"
                })
        
        return results

    def _compare_position_values(self, ai_value: str, sql_value: str) -> bool:
        """
        比较职位描述字段，支持多职位顺序不一致的情况
        
        Args:
            ai_value: AI提取的职位描述
            sql_value: SQL查询的职位描述
            
        Returns:
            是否匹配
        """
        ai_positions = set(p.strip() for p in ai_value.split("、") if p.strip())
        sql_positions = set(p.strip() for p in sql_value.split("、") if p.strip())
        return ai_positions == sql_positions

    def _compare_hold_fields(self, ai_data: Dict[str, Any], sql_data: Dict[str, Any]) -> str:
        """
        比对AI数据和SQL数据的持股字段，并返回格式化的比对结果
        
        Args:
            ai_data: AI提取的数据
            sql_data: SQL查询的数据
            
        Returns:
            格式化的比对结果字符串
        """
        error_messages = []

        for field_name in HOLD_FIELDS_TO_COMPARE:
            ai_value = ai_data.get(field_name, "")
            sql_value = sql_data.get(field_name, "")

            if field_name == "职位描述":
                if not self._compare_position_values(str(ai_value), str(sql_value)):
                    error_messages.append(f"{field_name}错误【正式库：{sql_value}，AI：{ai_value}】")
            else:
                processed_ai_value = self._preprocess_value(ai_value)
                processed_sql_value = self._preprocess_value(sql_value)

                if not self._compare_values(processed_ai_value, processed_sql_value):
                    error_messages.append(f"{field_name}错误【正式库：{sql_value}，AI：{ai_value}】")

        if not error_messages:
            return "数据一致"

        return "\n".join(error_messages)

    def _preprocess_value(self, value: Any) -> Any:
        """预处理值，统一格式"""
        if value is None:
            return ""

        str_value = str(value).strip()

        if not str_value:
            return ""

        if str_value in ("不适用", "-", "/", "－"):
            return ""

        str_value = str_value.replace(',','')

        is_negative = False
        if str_value.startswith('(') and str_value.endswith(')'):
            is_negative = True
            str_value = str_value[1:-1]

        if self._is_numeric_value(str_value):
            if '%' in str_value:
                str_value = str_value.replace("%", "")
            
            try:
                num_value = float(str_value)
                return -num_value if is_negative else num_value
            except ValueError:
                return str_value

        return str_value

    def _is_numeric_value(self, value: str) -> bool:
        """判断值是否为数值类型"""
        if not value:
            return False

        has_digit = any(c.isdigit() for c in value)

        numeric_chars = {'.', '-', '+', 'e', 'E', '%'}
        has_numeric_char = any(c in numeric_chars for c in value)

        return has_digit and (has_numeric_char or value.isdigit())

    def load_prompt_from_md(self, md_file_path: str = "prompt_ldrcg.md") -> str:
        """从MD文件加载提示词"""
        try:
            base_dir = get_base_path()
            abs_md_path = os.path.join(base_dir, md_file_path)

            if os.path.exists(abs_md_path):
                with open(abs_md_path, 'r', encoding='utf-8') as f:
                    return f.read()
            else:
                logger.warning(f"提示词文件不存在: {abs_md_path}")
                return ""
        except Exception as e:
            logger.error(f"加载提示词文件失败: {e}")
            return ""

    def _extract_data(self, file_id: str, filename: str) -> Optional[Dict[str, Any]]:
        """
        使用提示词提取数据
        
        Args:
            file_id: 上传后的文件ID
            filename: 文件名
            
        Returns:
            提取的数据，失败返回None
        """
        try:
            prompt = self.load_prompt_from_md(PROMPT_FILE)
            if not prompt:
                logger.error(f"加载提示词失败: {filename}")
                return None
            
            ai_data_results = enhanced_ai_service.extract_data_from_file(file_id, prompt)
            if ai_data_results:
                return ai_data_results
            else:
                logger.warning(f"AI数据提取返回空结果: {filename}")
                return None
                
        except Exception as e:
            logger.error(f"AI数据提取异常: {filename} - {e}")
            return None

    def generate_report(self, results: List[Dict[str, Any]], report_file: str = None) -> str:
        """生成比对报告"""
        if not results:
            print("没有可生成报告的数据")
            return ""

        base_dir = get_base_path()
        report_dir = os.path.join(base_dir, "report")
        if not os.path.exists(report_dir):
            os.makedirs(report_dir)
            print(f"创建报告目录: {report_dir}")

        if not report_file:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_file = os.path.join(report_dir, f"领导人持股报酬比对报告_{timestamp}.xlsx")

        try:
            with pd.ExcelWriter(report_file, engine='openpyxl') as writer:
                self._create_comparison_sheet(results, writer)

            print(f"报告已生成: {report_file}")
            return report_file

        except Exception as e:
            print(f"生成报告失败: {e}")
            return ""

    def _extract_display_name(self, pdf_file_name: str) -> str:
        """
        从PDF文件名中提取显示名称
        去除"股票代码-信息发布日期-"前缀
        
        Args:
            pdf_file_name: PDF文件名，格式如"000001-2025-03-15-公告标题.pdf"
            
        Returns:
            提取后的公告标题，如"公告标题.pdf"
        """
        if not pdf_file_name:
            return pdf_file_name
        
        base_name = os.path.splitext(pdf_file_name)[0]
        parts = base_name.split('-')
        
        if len(parts) >= 4:
            display_name = '-'.join(parts[4:])
            if display_name:
                return display_name + os.path.splitext(pdf_file_name)[1]
        
        return pdf_file_name

    def _create_comparison_sheet(self, results: List[Dict[str, Any]], writer: pd.ExcelWriter):
        """创建比对结果表"""
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        comparison_data = []

        for result in results:
            comparison_results = result.get("comparison_result", [])
            pdf_file_path = result.get("pdf_file_path", "")
            pdf_file_name = result.get("pdf_file_name", "")

            if not comparison_results:
                continue

            display_name = self._extract_display_name(pdf_file_name)
            
            sorted_results = []
            salary_result = None
            hold_results = []
            
            for comparison in comparison_results:
                ldxm = comparison.get("领导人姓名", "")
                if ldxm == "报酬比对":
                    salary_result = comparison
                else:
                    hold_results.append(comparison)
            
            if salary_result:
                sorted_results.append(salary_result)
            sorted_results.extend(hold_results)

            for comparison in sorted_results:
                comparison_data.append({
                    "公告标题": display_name,
                    "GPDM": comparison.get("GPDM", ""),
                    "XXFBRQ": comparison.get("XXFBRQ", ""),
                    "领导人姓名": comparison.get("领导人姓名", ""),
                    "比对结果": comparison.get("比对结果", ""),
                    "_pdf_path": pdf_file_path
                })

        df = pd.DataFrame(comparison_data)
        
        if df.empty:
            df.to_excel(writer, sheet_name="比对结果", index=False)
            return
        
        df.to_excel(writer, sheet_name="比对结果", index=False)
        
        workbook = writer.book
        worksheet = writer.sheets["比对结果"]
        
        hyperlink_font = Font(color="0563C1", underline="single")
        wrap_alignment = Alignment(wrap_text=True, vertical='top')
        
        for idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            pdf_path = row[5] if len(row) > 5 else ""
            comparison_result = row[4] if len(row) > 4 else ""
            
            if pdf_path:
                cell = worksheet.cell(row=idx, column=1)
                cell.hyperlink = pdf_path
                cell.font = hyperlink_font
            
            result_cell = worksheet.cell(row=idx, column=5)
            result_cell.alignment = wrap_alignment
            
            if comparison_result and "\n" in comparison_result:
                line_count = comparison_result.count("\n") + 1
                worksheet.row_dimensions[idx].height = max(13.5, line_count * 13.5)
        
        worksheet.column_dimensions['C'].width = 12
        worksheet.column_dimensions['E'].width = 80
        
        if worksheet.max_column >= 6:
            worksheet.delete_cols(6)


def main():
    """主函数"""
    print("=" * 60)
    print("领导人持股报酬比对系统")
    print("=" * 60)

    try:
        processor = LeaderStockSalaryProcessor()

        while True:
            print("\n请选择操作:")
            print("1. 处理指定目录文件")
            print("2. 退出")

            choice = input("\n请输入选项 (1-2): ").strip()

            if choice == "1":
                custom_dir = input("请输入要处理的目录路径: ").strip()
                if not custom_dir:
                    print("目录路径不能为空")
                    continue

                base_dir = Path(custom_dir)
                if not base_dir.exists():
                    print(f"目录不存在: {custom_dir}")
                    continue

                pdf_files = list(base_dir.glob("*.pdf"))
                if not pdf_files:
                    print(f"在目录 {custom_dir} 中未找到PDF文件")
                    continue

                program_start_time = datetime.now()
                logger.info(f"程序开始执行 - 目录: {custom_dir}, 文件数量: {len(pdf_files)}")
                print(f"\n开始处理，文件数量: {len(pdf_files)}...")

                start_time = datetime.now()
                results = processor.process_all_files(base_dir)
                end_time = datetime.now()

                print(f"\n\n处理完成! 共处理 {len(results)} 个文件，耗时: {end_time - start_time}")

                success_count = 0
                no_data_count = 0
                failed_files = []
                no_data_files = []
                with processor.lock:
                    for file_path, status in processor.file_status.items():
                        if status == "completed":
                            success_count += 1
                        elif status == "no_data":
                            no_data_count += 1
                            no_data_files.append(file_path.name)
                        else:
                            failed_files.append((file_path.name, status))
                
                print(f"成功处理: {success_count} 个文件")
                print(f"无披露数据: {no_data_count} 个文件")
                if no_data_files:
                    for file_name in no_data_files[:5]:
                        print(f"  - {file_name}")
                    if len(no_data_files) > 5:
                        print(f"  ... 还有 {len(no_data_files) - 5} 个文件无披露数据")
                if failed_files:
                    print(f"处理失败: {len(failed_files)} 个文件")
                    for file_name, status in failed_files[:5]:
                        print(f"  - {file_name}: {status}")
                    if len(failed_files) > 5:
                        print(f"  ... 还有 {len(failed_files) - 5} 个文件处理失败")

                if results:
                    print("\n生成处理报告...")
                    base_dir = get_base_path()
                    report_dir = os.path.join(base_dir, "report")
                    if not os.path.exists(report_dir):
                        os.makedirs(report_dir)
                        print(f"创建报告目录: {report_dir}")

                    report_file = os.path.join(report_dir,
                                               f"领导人持股报酬比对报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
                    processor.generate_report(results, report_file)

                    program_end_time = datetime.now()
                    total_duration = program_end_time - program_start_time
                    logger.info(f"程序执行完成 - 总耗时: {total_duration}")
                    print(f"\n程序执行完成! 总耗时: {total_duration}")
                else:
                    print("\n没有成功处理的文件，请检查日志获取详细信息")

            elif choice == "2":
                print("退出程序")
                break

            else:
                print("无效选项，请重新选择")

    except KeyboardInterrupt:
        print("\n用户中断操作")
    except Exception as e:
        logger.error(f"处理过程中发生错误: {e}", exc_info=True)
        print(f"程序运行出错: {e}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()
