from datetime import datetime

import pandas as pd
import pyodbc
import glob
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# Wind最新结果文件  从终端下载
wd_excel_path = "实际控制人-wd.xlsx"

SERVER = '10.102.25.11,8080'  # 服务器名称或IP地址
USERNAME = 'WebResourceNew_Read'  # 登录用户名
PASSWORD = 'New_45ted'  # 登录密码
DRIVER = 'ODBC Driver 17 for SQL Server'  # ODBC驱动版本

sql = '''/*最新背景表实际控制人提取*/
WITH AAA AS
(Select C.IGSDM,cast(C.XXFBRQ as date) XXFBRQ,cast(C.JZRQ as date) JZRQ,C.GDMC,C.GDXZ FROM  [10.101.0.212].JYPRIME.dbo.usrZYGDBJJS C
Where C.CZYF=1
AND C.GDXH=9),
/*实际控制人表数据*/
BBB AS(Select DISTINCT C.IGSDM,cast(C.XXFBRQ as date) XXFBRQ,cast(C.JZRQ as date) JZRQ,STUFF((SELECT ','+CAST(SJKZR AS VARCHAR(100)) /*调整处罚机构的格式*/
                              FROM  [10.101.0.212].JYPRIME.dbo.usrGSSJKZR
					          WHERE IGSDM=C.IGSDM
							  AND JZRQ=C.JZRQ
							  AND XXFBRQ=C.XXFBRQ
					          FOR XML PATH('')),1,1,'') AS SJKZR FROM  [10.101.0.212].JYPRIME.dbo.usrGSSJKZR C
Where Not Exists(Select 1 From  [10.101.0.212].JYPRIME.dbo.usrGSSJKZR Where IGSDM=C.IGSDM AND JZRQ>C.JZRQ)
GROUP BY C.IGSDM,C.ID,C.XXFBRQ,C.JZRQ,C.SJKZR)
select A.GPDM 股票代码,A.ZQJC 证券简称,B.MS 上市标志,C.MS 上市状态,AAA.XXFBRQ 信息发布日期1,AAA.JZRQ 截止日期1,AAA.GDMC JY背景表实际控制人,BBB.XXFBRQ 信息发布日期2,BBB.JZRQ 截止日期2,BBB.SJKZR JY实际控制人表
from  [10.101.0.212].JYPRIME.dbo.usrZQZB A
LEFT JOIN  [10.101.0.212].JYPRIME.dbo.usrXTCLB B ON B.LB=207 AND A.SSBZ=B.DM
LEFT JOIN  [10.101.0.212].JYPRIME.dbo.usrXTCLB C ON C.LB=1176 AND A.SSZT=C.DM
LEFT JOIN AAA ON AAA.IGSDM=A.IGSDM
LEFT JOIN BBB ON BBB.IGSDM=A.IGSDM
WHERE A.ZQSC In(83,90,18)
AND A.ZQLB IN(1,2)
AND (A.SSZT NOT IN(5,9) OR (A.SSBZ=7 AND A.SSZT=9) )
  AND A.GPDM NOT LIKE 'X_____'
ORDER BY 1'''


#读取excel文件
def read_excel():
    # 读取第一页（主要数据）和第二页（统计数据）
    excel_data = pd.read_excel(wd_excel_path, sheet_name=0, dtype=str).fillna('')
    excel_sheet2 = pd.read_excel(wd_excel_path, sheet_name=1, dtype=str).fillna('')
    # 加载原始工作簿以保持格式
    wb_source = load_workbook(wd_excel_path)
    return excel_data, excel_sheet2, wb_source

#查询数据库
def read_sql():
    conn = pyodbc.connect(SERVER=SERVER, UID=USERNAME, PWD=PASSWORD, DRIVER=DRIVER)
    cursor = conn.cursor().execute(sql)
    columns = [column[0] for column in cursor.description]

    sql_data = []
    for row in cursor.fetchall():
        sql_data.append(dict(zip(columns, row)))

    conn.close()
    return sql_data

def normalize_controller_names(controller_str):
    """
    标准化实际控制人名称，处理多个实际控制人的情况
    :param controller_str: 实际控制人名称字符串
    :return: 标准化后的实际控制人名称集合
    """
    if not controller_str or controller_str == '' or pd.isna(controller_str):
        return set()

    # 使用正则表达式根据逗号和顿号分隔实际控制人名称
    # 同时处理可能的空格
    controllers = re.split(r'[、,，\s]+', controller_str.strip())

    # 过滤掉空字符串并去除两端空格
    controllers = [ctrl.strip() for ctrl in controllers if ctrl.strip()]

    # 返回集合以忽略顺序
    return set(controllers)

def preprocess_compare_data(excel_data, sql_data):
    """
    预处理比对数据，找出与最新比对结果文件中的差异
    :param excel_data: 当前Excel数据
    :param sql_data: 当前SQL数据
    :return: preprocess_excel_data, preprocess_sql_data - 与上次结果不同的数据
    """
    
    # 查找最新的比对结果文件
    pattern = "权益部-境内股票-增发组-实际控制人三方比对结果*.xlsx"
    result_files = glob.glob(pattern)
    
    if not result_files:
        print("未找到历史比对结果文件，返回空数据")
        return [], []
    
    # 提取日期并找到最新的文件
    latest_file = None
    latest_date = None
    
    for file in result_files:
        try:
            # 从文件名中提取日期部分
            date_str = file.split("权益部-境内股票-增发组-实际控制人三方比对结果(")[1].replace(").xlsx", "")
            if "年" in date_str and "月" in date_str and "日" in date_str:
                # 处理 YYYY年Y月D日 格式
                date_str = date_str.replace("年", "-").replace("月", "-").replace("日", "")
                file_date = datetime.strptime(date_str, "%Y-%m-%d")
            
            if latest_date is None or file_date > latest_date:
                latest_date = file_date
                latest_file = file
        except:
            # 跳过无法解析日期的文件
            continue
    
    if latest_file is None:
        print("未找到有效的历史比对结果文件，返回空数据")
        return [], []
    
    print(f"找到最新比对结果文件: {latest_file}")
    
    # 读取最新比对结果文件中的WD和JY sheet
    try:
        # 读取WD sheet
        wd_data = pd.read_excel(latest_file, sheet_name='WD', dtype=str).fillna('')
        # 读取JY sheet，确保股票代码保留前导零
        def fix_stock_code(x):
            if pd.isna(x) or x == '':
                return ''
            stock_code = str(x).replace('.0', '').strip()
            if stock_code.isdigit() and len(stock_code) < 6:
                return stock_code.zfill(6)
            return stock_code
        
        jy_data = pd.read_excel(latest_file, sheet_name='JY', converters={
            '股票代码': fix_stock_code
        }).fillna('')
    except Exception as e:
        print(f"读取历史比对结果文件出错: {e}")
        return [], []
    
    # 创建结果存储
    preprocess_excel_data = []
    preprocess_sql_data = []
    
    # 将历史数据转换为字典，便于比对
    wd_dict = {}
    for _, row in wd_data.iterrows():
        stock_code = row['证券代码']
        wd_dict[stock_code] = row
    
    jy_dict = {}
    for _, row in jy_data.iterrows():
        stock_code = row['股票代码']
        jy_dict[stock_code] = row
    
    # 比对当前Excel数据与历史WD数据
    for _, row in excel_data.iterrows():
        stock_code = row['证券代码']
        current_controller = row['实际控制人名称[日期] 最新']
        
        # 如果历史数据中存在该股票代码，且实际控制人不同，则添加到结果中
        if stock_code in wd_dict:
            historical_controller = wd_dict[stock_code]['实际控制人名称[日期] 最新']
            # 标准化实际控制人名称以便比对
            current_set = normalize_controller_names(current_controller)
            historical_set = normalize_controller_names(historical_controller)
            
            if current_set != historical_set:
                preprocess_excel_data.append(row)
        else:
            # 如果历史数据中不存在该股票代码，也添加到结果中
            preprocess_excel_data.append(row)
    
    # 比对当前SQL数据与历史JY数据
    for row in sql_data:
        stock_code = row['股票代码']
        current_background = row['JY背景表实际控制人']
        current_actual = row['JY实际控制人表']
        
        # 如果历史数据中存在该股票代码，且实际控制人不同，则添加到结果中
        if stock_code in jy_dict:
            historical_background = jy_dict[stock_code]['JY背景表实际控制人']
            historical_actual = jy_dict[stock_code]['JY实际控制人表']
            
            # 标准化实际控制人名称以便比对
            current_background_set = normalize_controller_names(current_background)
            historical_background_set = normalize_controller_names(historical_background)
            current_actual_set = normalize_controller_names(current_actual)
            historical_actual_set = normalize_controller_names(historical_actual)
            
            # 只要有一列值不同，就保存到结果中
            if (current_background_set != historical_background_set or 
                current_actual_set != historical_actual_set):
                preprocess_sql_data.append(row)

        else:
            # 如果历史数据中不存在该股票代码，也添加到结果中
            preprocess_sql_data.append(row)

    
    return preprocess_excel_data, preprocess_sql_data


def compare_data(preprocess_excel_data, preprocess_sql_data):
    """
    比对Excel数据和SQL数据
    :param excel_data: Excel数据
    :param sql_data: SQL数据
    :return: 比对结果
    """
    # 将SQL数据转换为以股票代码前6位为键的字典，方便查找
    sql_dict = {}
    # 检查sql_data是否为DataFrame，如果是则使用iterrows，否则直接遍历列表
    if hasattr(preprocess_sql_data, 'iterrows'):
        # DataFrame情况
        for _, row in preprocess_sql_data.iterrows():
            stock_code_prefix = row['股票代码']
            sql_dict[stock_code_prefix] = row
    else:
        # 列表情况
        for row in preprocess_sql_data:
            stock_code_prefix = row['股票代码']
            sql_dict[stock_code_prefix] = row
    
    # 将Excel数据转换为以股票代码前6位为键的字典，方便查找
    excel_dict = {}
    # 检查excel_data是否为DataFrame，如果是则使用iterrows，否则直接遍历列表
    if hasattr(preprocess_excel_data, 'iterrows'):
        # DataFrame情况
        for _, row in preprocess_excel_data.iterrows():
            stock_code = row['证券代码']
            stock_code_prefix = stock_code[:6] if len(stock_code)>6 else stock_code
            excel_dict[stock_code_prefix] = row
    else:
        # 列表情况
        for row in preprocess_excel_data:
            stock_code = row['证券代码']
            stock_code_prefix = stock_code[:6] if len(stock_code)>6 else stock_code
            excel_dict[stock_code_prefix] = row

    # 创建比对结果列表
    comparison_results = []

    # 第一部分：遍历Excel数据进行比对（Excel中有但SQL中没有或不匹配的数据）
    for row in preprocess_excel_data:
        stock_code = row['证券代码']
        # 使用前6位数字进行比对
        stock_code_prefix = stock_code[:6] if len(stock_code)>6 else stock_code
        excel_controller = row['实际控制人名称[日期] 最新']

        # 初始化比对结果
        result = ""
        
        # 检查SQL中是否存在该股票代码的数据
        if stock_code_prefix in sql_dict:
            sql_row = sql_dict[stock_code_prefix]
            background_controller = str(sql_row['JY背景表实际控制人']).replace('无实际控制人','') if sql_row['JY背景表实际控制人'] is not None else ''
            actual_controller = str(sql_row['JY实际控制人表']).replace('无实际控制人','') if sql_row['JY实际控制人表'] is not None else ''
            
            # 标准化实际控制人名称以便比对
            excel_controller_set = normalize_controller_names(excel_controller)
            background_controller_set = normalize_controller_names(background_controller)
            actual_controller_set = normalize_controller_names(actual_controller)
            
            # 如果Excel中的值与SQL两个表的值都不一致
            if excel_controller_set != background_controller_set and excel_controller_set != actual_controller_set:
                result = "比对不一致"

            # 单个表不一致
            elif excel_controller_set != background_controller_set and len(background_controller_set) > 0:
                result = "背景表比对不一致"

            elif excel_controller_set != actual_controller_set and len(actual_controller_set) > 0:
                result = "实际控制人表比对不一致"

            comparison_results.append({
            '证券代码': stock_code,
            '证券简称': row['证券简称'],
            '实际控制人名称[日期] 最新': excel_controller,
            '股票代码': sql_row.get('股票代码', ''),
            '股票简称': sql_row.get('证券简称', ''),
            '上市标志': sql_row.get('上市标志', ''),
            '上市状态': sql_row.get('上市状态', ''),
            '信息发布日期1': sql_row.get('信息发布日期1', ''),
            '截止日期1': sql_row.get('截止日期1', ''),
            'JY背景表实际控制人': sql_row.get('JY背景表实际控制人', ''),
            '信息发布日期2': sql_row.get('信息发布日期2', ''),
            '截止日期2': sql_row.get('截止日期2', ''),
            'JY实际控制人表': sql_row.get('JY实际控制人表', ''),
            '变动源': 'WD/JY变动',
            '差异性说明': '',
            '是否已修改': '',
            '未修改原因': '',
            '比对结果': result
            })

        else:
            comparison_results.append({
            '证券代码': stock_code,
            '证券简称': row['证券简称'],
            '实际控制人名称[日期] 最新': excel_controller,
            '股票代码': '',
            '股票简称': '',
            '上市标志': '',
            '上市状态': '',
            '信息发布日期1': '',
            '截止日期1': '',
            'JY背景表实际控制人': '',
            '信息发布日期2': '',
            '截止日期2': '',
            'JY实际控制人表': '',
            '变动源': 'WD变动',
            '差异性说明': '',
            '是否已修改': '',
            '未修改原因': '',
            '比对结果': ''
            })
        
    
    # 第二部分：遍历SQL数据，找出SQL中有但Excel中没有的数据
    for row in preprocess_sql_data:
        stock_code = row['股票代码']
        # 使用前6位数字进行比对
        stock_code_prefix = stock_code[:6] if len(stock_code)>6 else stock_code
        
        # 检查Excel中是否存在该股票代码的数据
        if stock_code_prefix not in excel_dict:
            # SQL中有但Excel中没有的数据
            background_controller = str(row['JY背景表实际控制人']).replace('无实际控制人','') if row['JY背景表实际控制人'] is not None else ''
            actual_controller = str(row['JY实际控制人表']).replace('无实际控制人','') if row['JY实际控制人表'] is not None else ''
            
            comparison_results.append({
                '证券代码': '',
                '证券简称': '',
                '实际控制人名称[日期] 最新': '',
                '股票代码': row.get('股票代码', ''),
                '股票简称': row.get('证券简称', ''),
                '上市标志': row.get('上市标志', ''),
                '上市状态': row.get('上市状态', ''),
                '信息发布日期1': row.get('信息发布日期1', ''),
                '截止日期1': row.get('截止日期1', ''),
                'JY背景表实际控制人': background_controller,
                '信息发布日期2': row.get('信息发布日期2', ''),
                '截止日期2': row.get('截止日期2', ''),
                'JY实际控制人表': actual_controller,
                '变动源': 'JY变动',
                '差异性说明': '',
                '是否已修改': '',
                '未修改原因': '',
                '比对结果': ''
            })

    return comparison_results

def save_results(comparison_results, excel_data, excel_sheet2, sql_data, wb_source):
    """
    保存比对结果到Excel文件
    :param comparison_results: 比对结果
    :param excel_data: 原始Excel数据
    :param excel_sheet2: Excel第二页数据
    :param sql_data: SQL数据
    :param wb_source: 原始Excel工作簿对象
    """
    now_date = datetime.now().strftime('%Y年%m月%d日')
    
    # 创建新的工作簿
    wb_new = Workbook()
    # 删除默认工作表
    wb_new.remove(wb_new.active)
    
    # 原样复制第二页数据到"比对结果统计" sheet
    if len(wb_source.worksheets) > 1:
        # 获取原始文件的第二页（索引为1）
        source_sheet = wb_source.worksheets[1]
        
        # 创建目标工作表
        target_sheet = wb_new.create_sheet("比对结果统计")
        
        # 复制所有数据
        for row in source_sheet.iter_rows():
            for cell in row:
                # 复制值
                new_cell = target_sheet.cell(row=cell.row, column=cell.column)
                new_cell.value = cell.value
                
                # 复制格式
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = copy(cell.number_format)
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # 复制列宽（在数据复制完成后进行）
        for col_letter in source_sheet.column_dimensions:
            source_col_dim = source_sheet.column_dimensions[col_letter]
            target_col_dim = target_sheet.column_dimensions[col_letter]
            if source_col_dim.width:
                target_col_dim.width = source_col_dim.width
        
        # 复制行高
        for row_idx in source_sheet.row_dimensions:
            source_row_dim = source_sheet.row_dimensions[row_idx]
            target_row_dim = target_sheet.row_dimensions[row_idx]
            if source_row_dim.height:
                target_row_dim.height = source_row_dim.height
    
    # 保存比对结果到"比对结果" sheet
    if not comparison_results:
        comparison_df = pd.DataFrame(columns=[
            '证券代码', '证券简称', '实际控制人名称[日期] 最新', 
            '股票代码', '股票简称', '上市标志', '上市状态',
            '信息发布日期1', '截止日期1', 'JY背景表实际控制人',
            '信息发布日期2', '截止日期2', 'JY实际控制人表',
            '变动源', '差异性说明', '是否已修改', '未修改原因', '比对结果'
        ])
    else:
        comparison_df = pd.DataFrame(comparison_results)
    
    # 添加比对结果工作表
    ws_comparison = wb_new.create_sheet("比对结果")
    for r in dataframe_to_rows(comparison_df, index=False, header=True):
        ws_comparison.append(r)
    
    # 保存原始Excel数据到"WD" sheet
    ws_wd = wb_new.create_sheet("WD")
    for r in dataframe_to_rows(excel_data, index=False, header=True):
        ws_wd.append(r)
    
    # 保存SQL数据到"JY" sheet
    sql_df = pd.DataFrame(sql_data)
    ws_jy = wb_new.create_sheet("JY")
    for r in dataframe_to_rows(sql_df, index=False, header=True):
        ws_jy.append(r)
    
    # 保存文件
    wb_new.save(f'权益部-境内股票-增发组-实际控制人三方比对结果({now_date}).xlsx')

def run_comparison():
    # 读取Excel数据
    excel_data, excel_sheet2, wb_source = read_excel()
    print(f"Excel数据读取完成，共{len(excel_data)}条记录")
    
    # 读取SQL数据
    print("正在查询数据库……")
    start_time = datetime.now()
    sql_data = read_sql()
    end_time = datetime.now()
    use_time = end_time-start_time
    print(f"SQL数据读取完成，共{len(sql_data)}条记录，耗时{use_time}")
    
    # 预处理比对数据，找出与上次结果的差异
    print("开始预处理比对数据...")
    preprocess_excel_data, preprocess_sql_data = preprocess_compare_data(excel_data, sql_data)

    # 进行数据比对
    comparison_results = compare_data(preprocess_excel_data, preprocess_sql_data)
    print(f"数据比对完成，共{len(comparison_results)}条记录")
    
    # 保存结果
    save_results(comparison_results, excel_data, excel_sheet2, sql_data, wb_source)
    print("比对结果已保存！")

def main():
    """主循环，等待用户输入并执行比对"""
    print("实际控制人三方比对程序启动")
    print("按回车键开始运行，输入'quit'退出程序")
    
    while True:
        try:
            # 等待用户输入
            user_input = input("\n是否开始运行？")
            
            # 检查用户是否要退出
            if user_input.lower().strip() == 'quit':
                print("程序已退出")
                break
            
            # 执行比对流程
            run_comparison()
            
        except KeyboardInterrupt:
            print("\n\n程序被用户中断")
            break
        except Exception as e:
            print(f"\n程序执行出错: {e}")
            print("程序将继续运行，如需退出请输入'quit'")

if __name__ == "__main__":
    main()



