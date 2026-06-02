"""
领导人数据比对模块 - 申报稿领导人AI提取数据与数据库已有数据比对
迁移自 领导人介绍-日常比对/main.py 的比对逻辑
"""
import difflib
import os
from typing import Any, Dict, List, Optional, Tuple, Union

import pandas as pd

from database_manager import db_manager
from logger_config import get_logger
from path_utils import get_reports_dir

logger = get_logger(__name__)

SQL_QUERY = '''
SELECT A.ID,B.GPDM,A.XM,CASE WHEN A.XB=1 THEN '男' WHEN A.XB=2 THEN '女' ELSE '' END AS XB,
       A.CSRQ,C.MS XL,D.MS GJ,E.ZWMC,E.ZW,CASE WHEN E.CZYF=1 THEN '在任' ELSE '离任' END AS CZYF,BJJS
FROM [10.101.0.212].JYPRIME.dbo.usrGSZYLDRJS A
    JOIN [10.101.0.212].JYPRIME.dbo.usrZQZB B
     ON A.INBBM=B.INBBM AND B.ZQSC IN (18,83,90) AND B.ZQLB IN (1,2,41)
    LEFT JOIN [10.101.0.212].JYPRIME.dbo.usrXTCLB C
     ON A.XL=C.DM AND C.LB=1154
    LEFT JOIN [10.101.0.212].JYPRIME.dbo.usrXTCLB D
     ON A.GJ=D.DM AND D.LB=1023
    JOIN (
        SELECT ZWMC,ZW,CZYF,RID,
               ROW_NUMBER() OVER (PARTITION BY RID, ZWMC, ZW ORDER BY RZQSR DESC) as rn
        FROM [10.101.0.212].JYPRIME.dbo.usrGSZYLDRJSRZQK
    ) E
     ON A.ID=E.RID AND E.rn = 1
WHERE B.GPDM = ? AND A.XM = ?
'''

FIELD_MAPPING = {
    "领导人姓名": "XM",
    "性别": "XB",
    "出生日期": "CSRQ",
    "学历": "XL",
    "国籍": "GJ",
}

POSITION_MAPPING = {
    "CFO": "15", "CTO": "27", "内控审计中心负责人": "47", "审计部副部长": "47",
    "审计监察中心负责人": "47", "审计经理": "47", "内部审计部负责人": "47",
    "内部审计机构负责人": "47", "内审部主任": "47", "内审机构负责人": "47",
    "审计部审计负责人": "47", "审计部主任": "47", "审计合规部负责人": "47",
    "审计责任人": "47", "财务部部长": "15", "财务负责人（CFO）": "15",
    "内部控制与审计部负责人": "47", "内部审计机构（审计监督部）负责人": "47",
    "内部审计主管": "47", "内审部负责人": "47", "内审部经理": "47",
    "内审部门负责人": "47", "审计部部长": "47", "审计部负责人": "47",
    "审计处负责人": "47", "审计法律部部长": "47", "代理财务负责人": "303",
    "代理董事局主席": "46", "审计法务部部长": "47", "审计风控部负责人": "47",
    "财务负责人": "15", "代理内部审计负责人": "47", "审计监察总监": "47",
    "代理总法律顾问": "61", "党委书记": "210", "党委委员": "213",
    "风控审计本部部长(内部审计负责人)": "47", "内控审计部部长": "47",
    "内审部总监": "47", "内审负责人": "47", "审计部经理": "47",
    "审计部门负责人": "47", "审计中心副总经理": "47", "非独立外部董事": "3",
    "监察审计部负责人": "47", "内部审部负责人": "47", "内部审计部门负责人": "47",
    "风险管理委员会委员": "605", "内审监察部负责人": "47", "审计监察部负责人": "47",
    "总审计师(审计责任人)": "47", "内部审计部经理": "47", "风险控制委员会成员": "605",
    "风险控制委员会召集人": "605", "内部审计负责人": "47", "审计部": "47",
    "代理合规总监": "84", "审计部门（审计法律部）负责人": "47", "审计负责人": "47",
    "副总裁(首席产品官)": "12", "副总经理(副总裁)": "12", "内控内审部负责人": "47",
    "内控审计经理": "47", "内审室负责人": "47", "股东监事": "8",
    "审计部总监": "47", "审计部总经理": "47", "代理首席风险官": "60",
    "关联交易控制委员会主席委员": "606", "关联交易审核委员会委员": "606",
    "审计部总经理（审计机构负责人）": "47", "审计机构负责人": "47",
    "审计室主任": "47", "审计部副主任": "47", "审计法务办公室总经理": "47",
    "审计中心负责人": "47", "审计总监": "47", "技术负责人": "27",
    "总审计师": "47", "审核委员会主任委员": "601", "审计管理委员会委员": "601",
    "计财部门负责人": "15", "审计及预算审核委员会成员": "601", "纪委书记": "82",
    "董事会办公室主任": "78", "审计委员会委员会成员": "601",
    "审计委员会主任(召集人)": "601", "审计委员会主席委员": "601",
    "财务与审计委员会主任委员": "601", "审计委员会副主席": "601", "轮值CEO": "45",
    "审计委员会负责人": "601", "审计委员会主任": "601",
    "审计委员会主任委员": "601", "审计与审核委员会召集人": "601",
    "财务审计委员会委员": "601", "董事会审计委员会委员": "601",
    "审计委员会副主任委员": "601", "审计委员会委主任委员": "601",
    "审计委员会召集人": "601", "审核委员会成员": "601",
    "全面风险管理委员会委员": "605", "财务审计委员会召集人": "601",
    "财务与审计委员会委员": "601", "独立审核委员会成员": "601",
    "审计委员会成员主任委员": "601", "审计委员会主任委员(召集人)": "601",
    "审计委员会主任委员（召集人）": "601", "财务/审计委员会主任委员": "601",
    "代理审计委员会委员": "601", "独立审核委员会主席": "601",
    "审核委员会委员": "601", "审核委员会召集人": "601",
    "审计委员主任委员": "601", "审计与审核委员会委员": "601",
    "审计委员会副主任": "601", "审计委员会主席": "601",
    "财务/审计委员会委员": "601", "独立审核委员会委员": "601",
    "审计委员会成员": "601", "审计委员会委员": "601",
    "审计委员会委员召集人": "601", "审计委员会委员主任委员": "601",
    "财务与审计委员会成员": "601", "代理审计委员会成员": "601",
    "审核委员会主席": "601", "审计/内控委员会委员": "601",
    "审计委员会委员主任": "601", "薪酬和考核委员会委员": "602",
    "薪酬和考核委员会召集人": "602", "薪酬委员会委员": "602",
    "薪酬与考核委员会往后": "602", "首席人力资源执行官": "30",
    "薪酬与考核委员会委员（召集人）": "602", "薪酬与考核委员会主任": "602",
    "薪酬与考核委员会主任委员（召集人）": "602", "代理薪酬与考核委员会委员": "602",
    "薪酬和考核委员会主任委员": "602", "提名委员会成员委员": "604",
    "提名委员会委员": "604", "提名委员会问委员": "604", "提名委员会主任": "604",
    "提名委员会主任委员": "604", "薪酬考核委员会召集人": "602",
    "薪酬与考核计委员会主任委员": "602", "薪酬与考核委员会成员": "602",
    "代理薪酬与考核委员会成员": "602", "考核和薪酬委员会委员": "602",
    "薪酬/考核委员会委员": "602", "薪酬/考核委员会主任委员": "602",
    "薪酬考核委员会成员": "602", "薪酬委员会主席": "602",
    "薪酬与考核委员会委员": "602", "薪酬与考核委员会委员主任": "602",
    "薪酬与考核委员召集人": "602", "薪酬与考评委员会成员": "602",
    "薪酬与考委员会委员": "602", "酬与考核委员会主任委员": "602",
    "薪酬委员会召集人": "602", "薪酬与绩效管理委员会委员": "602",
    "薪酬与考核委员会委员主任委员": "602", "薪酬与考核委员会主席": "602",
    "预算薪酬与考核委员会主任委员": "602", "薪酬与考核会委员": "602",
    "薪酬与考核计委员会委员": "602", "薪酬与考核委员会副主任委员": "602",
    "薪酬与考核委员会委委员": "602", "薪酬与考核委员会主任委员": "602",
    "薪酬与考核委员会主任委员(召集人)": "602", "薪酬与考核委员主任委员": "602",
    "预算薪酬与考核委员会委员": "602", "薪酬及考核委员会委员": "602",
    "薪酬考核委员会主任委员": "602", "薪酬委员会主任委员": "602",
    "薪酬与考核委员会副主任": "602", "考核与薪酬委员会主任委员": "602",
    "薪酬及考核委员会主任委员": "602", "薪酬考核和考核委员会委员": "602",
    "薪酬考核和考核委员会召集人": "602", "薪酬考核委员会委员": "602",
    "薪酬委员会成员": "602", "薪酬与考核委员会召集人": "602",
    "薪酬与绩效考核委员会委员": "602", "薪酬与绩效考核委员会主任委员": "602",
    "薪酬与考核委员会委员召集人": "602", "薪酬与考评委员会委员": "602",
    "发展战略规划委员会委员": "603", "发展战略委员会委员": "603",
    "略委员会委员": "603", "战略发展及投资委员会副主席": "603",
    "战略和投资委员会委员": "603", "战略及投资委员会主任委员": "603",
    "战略决策委员会成员": "603", "战略投资委员会副主任委员": "603",
    "战略委员会成员": "603", "法务总监": "61", "发展战略委员会召集人": "603",
    "战略发展委员会成员": "603", "财务副总监": "18",
    "战略发展委员会召集人": "603", "战略管理委员会召集人": "603",
    "财务总监(财务负责人)": "15", "战略及投资委员会召集人": "603",
    "安全与战略委员会委员": "603", "常务副总裁": "11",
    "代理董战略委员会召集人": "603", "代理首席财务官": "303",
    "代理首席合规官": "84", "代理首席执行官": "45",
    "代理首席执行官(CEO)": "45", "发展与战略委员会委员": "603",
    "投资与战略委员会委员": "603", "战略/投资发展委员会委员": "603",
    "战略管理委员会委员": "603", "战略规划委员会委员": "603",
    "董事局主席": "1", "战略和发展委员会委员": "603",
    "独立非执行董事": "4", "战略和投资委员会召集人": "603",
    "战略决策委员会召集人": "603", "战略决策委员会主任委员": "603",
    "战略投资委员会成员": "603", "战略投资委员会委员": "603",
    "投资决策及战略发展委员会成员": "603", "投资与战略委员会召集人": "603",
    "战略发展及投资委员会主席": "603", "战略发展委员会主任委员": "603",
    "副总经理（副总裁）": "12", "高级副总裁": "12",
    "战略规划委员会主任委员": "603", "战略和发展委员会召集人": "603",
    "战略和投资委员会成员": "603", "战略决策委员会委员": "603",
    "战略委员会负责人": "603", "战略委员会委员召集人": "603",
    "战略委员会召集人": "603", "战略与发展委员会成员": "603",
    "安全与战略委员会主任委员": "603", "核心技术人员": "40",
    "代理战略与ESG委员会召集人": "603", "发展战略委员会主任委员": "603",
    "投资战略委员会委员": "603", "战略/投资发展委员会主任委员": "603",
    "战略（投资决策）委员会委员": "603", "战略规划与ESG委员会成员": "603",
    "战略投资委员会召集人": "603", "战略委员会委员副主任委员": "603",
    "战略与发展委员会召集人": "603", "战略与发展委员会主任委员": "603",
    "发展战略与投资委员会主任委员": "603", "战略管理委员会成员": "603",
    "战略和投资委员会主任委员": "603", "战略决策委员成员": "603",
    "战略委员会委员主任委员": "603", "监事副主席": "7",
    "监事会临时召集人": "301", "战略与决策委员会委员": "603",
    "战略与决策委员会召集人": "603", "战略与投资管理委员会成员": "603",
    "战略与投资管理委员会主任委员": "603", "经营管理执行委员会副主任委员": "600",
    "战略与投资委员会召集人": "603", "战略委员会主任委员": "603",
    "战略与规划委员会召集人": "603", "战略与投资决策委员会成员": "603",
    "战略与投资委员会成员": "603", "战略与投资委员会主席": "603",
    "联席公司秘书": "55", "代理战略和投资委员会主任委员": "603",
    "代理战略委员会召集人": "603", "代理战略委员会主任委员": "603",
    "技术与发展战略委员会成员": "603", "投资决策及战略发展委员会召集人": "603",
    "内控与风险管理委员会主任委员": "605", "战略发展及投资委员会召集人": "603",
    "战略发展委员会主席": "603", "战略和发展委员会主任委员": "603",
    "战略委员会副主任委员": "603", "战略委员会委员": "603",
    "战略与投资委员会副主任": "603", "战略委员会主任委员(召集人)": "603",
    "战略与发展管理委员会主任委员": "603", "战略与发展委员会委员": "603",
    "战略与决策委员会主任委员": "603", "代理战略发展委员会主任委员": "603",
    "发展与战略委员会成员": "603", "战略（投资决策）委员会主任委员": "603",
    "战略发展及投资委员会委员": "603", "战略发展委员会委员": "603",
    "战略发展委员会主任": "603", "战略管理委员会主任委员": "603",
    "战略委员会副主任": "603", "战略委员会主任": "603",
    "战略委员会主席": "603", "战略与投资委员会副主任委员": "603",
    "战略预可持续委员会主任委员": "603", "首席风险官": "60",
    "首席运营专家": "62", "提名委员会成员": "604",
    "提名委员会副主任委员": "604", "提名委员会主任委员（召集人）": "604",
    "提名委员召集人": "604", "外部非独立董事": "3", "运营中心总监": "62",
    "监事会召集人": "6", "财务长": "15", "财务总裁(财务负责人)": "15",
    "财务总监（总会计师）": "15", "代理CEO": "45", "代理首席信息官": "59",
    "法律合规部负责人": "84", "非执行董事": "3",
    "风险管理委员会主任委员": "605", "风险管理委员会主席": "605",
    "副总裁": "12", "副总经理(会计机构负责人)": "12",
    "关联交易决策委员会委员": "606", "关联交易控制委员会召集人": "606",
    "关联交易控制委员会主任委员": "606", "核心员工": "40", "监事": "8",
    "经理": "10", "联席总经理": "306", "临时召集人": "301",
    "轮值总经理": "10", "内控与风险管理委员会委员": "605",
    "人工智能首席技术官": "27", "人力行政总监": "30", "人事部部长": "30",
    "荣誉董事长": "16", "市场总监": "29", "首席财务长": "15",
    "首席财务官(暨财务负责人)": "15", "首席风险合规官": "60",
    "首席信息官（总经理助理）": "59", "首席执行官": "45",
    "首席执行官（CEO）": "45", "提名委员会会召集人": "604",
    "提名委员会委员召集人": "604", "提名委员会主席": "604",
    "业务总监(公司与机构业务)": "70", "CEO": "45", "COO（首席运营官）": "62",
    "财务管理中心总监": "15", "财务管理中心总经理": "15", "代理董事长": "46",
    "代理证券事务代表": "305", "董事局副主席": "2", "独立董事": "4",
    "风控委员会成员": "605", "风控委员会主席": "605", "风控总监": "60",
    "副总": "12", "副总裁(副总经理)": "12", "副总经理(常务副总裁)": "11",
    "副总经理(职业经理人)": "12", "合规总监": "84", "技术部部长": "27",
    "经营管理执行委员会委员": "600", "经营管理执行委员会主任委员": "600",
    "名誉主席": "16", "人力行政部总监": "30", "首席执行官(CEO)": "45",
    "首席执行官(CEO、总经理)": "45", "授权代表": "55",
    "提名委员会委员主任": "604", "条例获授权代表": "55",
    "投资与风险管理委员会主任委员": "605", "外部监事": "9", "行长助理": "125",
    "首席合规官": "84", "首席运营官": "62", "财务部门负责人": "15",
    "财务负责人（财务总监）": "15", "代理董事会秘书": "304",
    "代理提名委员会委员": "604", "代理行长": "101", "代理总会计师": "303",
    "董事长": "1", "董事会秘书": "5", "风控委员会委员": "605",
    "风险控制委员会主任委员": "605", "风险控制委员会主席": "605",
    "副董事长": "2", "副总裁（副总经理）": "12", "股东代表监事": "8",
    "关联交易委员会委员": "606", "关联交易委员会主任委员": "606",
    "合规负责人": "84", "监事长": "6", "监事会副主席": "7",
    "监事会监事长": "6", "经理助理": "13", "联席董事长": "300",
    "人力资源部总经理": "30", "人事总监": "30", "首席财务官（CFO）": "15",
    "首席合规官（合规负责人）": "84", "首席技术官": "27", "首席信息官": "59",
    "首席运营官(COO)": "62", "首席运营官（COO）": "62", "外部董事": "4",
    "信息披露事务负责人": "85", "债务融资工具信息披露事务负责人": "85",
    "执行委员会成员": "600", "执行总经理": "73", "助理总裁": "13",
    "安全技术总监": "27", "财务副总监(财务负责人)": "18",
    "财务负责人(财务总监)": "15", "财务总监（财务负责人）": "15",
    "业务总裁": "73", "常务副总经理": "11", "代理合规负责人": "84",
    "代理监事长": "301", "代理监事会召集人": "301", "代理总经理": "302",
    "党委专职副书记": "81", "董事": "3", "董事会主席": "1",
    "法律事务部总监": "61", "风控委员会主任委员": "605",
    "风险管理委员会召集人": "605", "风险与资本管理委员会主任委员": "605",
    "副监事长": "7", "副经理": "12", "工会主席": "25",
    "关联交易委员会召集人": "606", "技术总监": "27",
    "人力资源部总监": "30", "人力资源总监(总经理助理)": "30", "业务总监": "70",
    "首席财务官": "15", "首席财务官（财务负责人）": "15", "运营总监": "62",
    "首席技术官(CTO)": "27", "提名委员主任委员": "604", "证券事务代表": "38",
    "执行委员会副主席": "600", "执行总裁": "73",
    "总法律顾问(首席合规官CCO)": "61", "总经理": "10",
    "执行委员会委员": "600", "总工程师(副总经理)": "35",
    "总经理(轮值)": "10", "备任授权代表": "55", "财务部负责人": "15",
    "财务总监": "15", "代理财务负责人（财务总监）": "303",
    "代理财务总监": "303", "代理监事会主席": "301", "党委常委": "213",
    "独立监事": "9", "非独立董事": "3", "风险管理负责人": "60",
    "风险管理委员会成员": "605", "风险与合规管理委员会成员": "605",
    "风险政策委员会委员": "605", "风险总监": "60", "副行长": "120",
    "副总工程师": "41", "副总经理": "12", "副总经理（挂职）": "12",
    "高级副总裁(SVP)": "12", "公司秘书": "55",
    "关联交易审核委员会召集人": "606", "关联交易审核委员会主任委员": "606",
    "合规管理负责人": "84", "全面风险管理委员会主任委员": "605",
    "人力资源总监": "30", "人力总监": "30", "首席技术官（CTO）": "27",
    "首席人力资源官": "30", "首席市场官": "29",
    "提名委员会主任委员(召集人)": "604", "行长": "101",
    "研发高级副总经理": "12", "执行委员会召集人": "600", "资深副总裁": "12",
    "总裁（总经理）": "10", "总会计师": "15", "总经理助理": "13",
    "执行副总裁": "74", "总裁": "10", "总经理(总裁)": "10",
    "职工代表董事": "49", "职工代表监事": "51", "职工监事": "51",
    "终身名誉董事长": "16", "总会计师（财务负责人）": "15",
    "总经理（总裁）": "10", "执委会委员": "600", "执行董事": "3",
    "执行副总经理": "74", "执行委员会主席": "600", "总法律顾问": "61",
    "总法律顾问(首席合规官)": "61", "总法律顾问（首席合规官 CCO）": "61",
    "总法律顾问（首席合规官）": "61", "总工程师": "35",
    "总会计师(财务总监)": "15", "办公室副主任": "79", "办公室主任": "78",
    "财务负责人(总会计师)": "15", "常务副总": "11", "代理总裁": "302",
    "党委副书记": "81", "非执行独立董事": "4", "风险控制委员会委员": "605",
    "风险与控制委员会委员": "605", "风险与控制委员会主任委员": "605",
    "副总经理(高级管理人员)": "12", "高级副总经理": "12", "股东董事": "3",
    "关联交易管理委员会委员": "606", "关联交易控制委员会成员": "606",
    "关联交易控制委员会主任委员（召集人）": "606", "合规部门负责人": "84",
    "执行委员": "600", "监事会主席": "6", "联席总裁": "306",
    "轮值总裁": "10", "名誉董事长": "16", "其他核心人员": "40",
    "执行委员会副主任": "600", "执行委员会主任": "600",
    "提名委员会委委员": "604", "提名委员会召集人": "604",
    "提名委员会主席委员": "604", "香港联交所授权代表": "55",
    "职工董事": "49", "执行委员会主任委员": "600", "助理总经理": "13",
    "专职外部董事": "4", "总裁(总经理)": "10", "总裁助理": "13",
    "总会计师(财务负责人)": "15", "总经济师": "76",
}


def preprocess_text(text: str) -> str:
    """预处理文本，去除干扰因素"""
    if not text:
        return ""
    
    result = text
    
    result = result.replace(" ", "").replace("\t", "").replace("\n", "").replace("\r", "")
    
    result = result.upper()
    
    char_replacements = {
        "：": ":",
        "；": ";",
        "，": ",",
        """: '"',
        """: '"',
        "'": "'",
        "'": "'",
        "—": "-",
        "（": "(",
        "）": ")",
    }
    
    for old_char, new_char in char_replacements.items():
        result = result.replace(old_char, new_char)
    
    return result


def compare_texts_char_by_char(text1: str, text2: str) -> Tuple[List[Tuple[str, bool]], List[Tuple[str, bool]]]:
    """逐字比对两个文本，返回带标红标记的结果
    
    Returns:
        (text1_parts, text2_parts): 每个是 [(文字片段, 是否标红), ...] 的列表
    """
    if not text1 and not text2:
        return [], []
    
    if not text1:
        return [], [(text2, True)]
    
    if not text2:
        return [(text1, True)], []
    
    matcher = difflib.SequenceMatcher(None, text1, text2)
    
    text1_parts = []
    text2_parts = []
    
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            text1_parts.append((text1[i1:i2], False))
            text2_parts.append((text2[j1:j2], False))
        elif tag == 'replace':
            text1_parts.append((text1[i1:i2], True))
            text2_parts.append((text2[j1:j2], True))
        elif tag == 'delete':
            text1_parts.append((text1[i1:i2], True))
        elif tag == 'insert':
            text2_parts.append((text2[j1:j2], True))
    
    return text1_parts, text2_parts


class LeaderComparisonProcessor:
    """领导人数据比对处理器 - 申报稿版本"""

    def __init__(self):
        self.field_mapping = FIELD_MAPPING
        self.position_mapping = POSITION_MAPPING

    def compare_with_database(self, extracted_data: List[Dict[str, Any]],
                              stock_code: str, publish_date: str) -> List[Dict[str, Any]]:
        """将AI提取的领导人数据与数据库已有数据进行比对

        Args:
            extracted_data: AI提取的领导人数据列表（每条记录含职位描述等字段）
            stock_code: 股票代码
            publish_date: 信息发布日期

        Returns:
            比对结果列表
        """
        if not extracted_data:
            return []

        all_comparison_results = []

        for leader_data in extracted_data:
            leader_name = str(leader_data.get("领导人姓名", "")).strip()
            position_desc = str(leader_data.get("职位描述", "")).strip()

            if not leader_name:
                all_comparison_results.append(
                    self._create_comparison_result({}, leader_data, stock_code, "AI数据缺少领导人姓名")
                )
                continue

            sql_data_list = self._query_leader_from_db(stock_code, leader_name)

            positions = [p.strip() for p in position_desc.replace("、", ",").split(",") if p.strip()]

            if not positions:
                positions = [""]

            for position in positions:
                comparison_results = self._compare_single_position(
                    sql_data_list, leader_data, position, stock_code
                )
                all_comparison_results.extend(comparison_results)

        return all_comparison_results

    def _query_leader_from_db(self, stock_code: str, leader_name: str) -> List[Dict[str, Any]]:
        """从数据库查询领导人数据"""
        try:
            sql_data = db_manager.execute_query(SQL_QUERY, (stock_code, leader_name))

            if not sql_data:
                logger.info(f"数据库无对应记录: {stock_code} {leader_name}")
                return []

            results = []
            for record in sql_data:
                result = {}
                for key, value in record.items():
                    result[key] = value if value is not None else ""
                results.append(result)

            logger.info(f"数据库查询到 {len(results)} 条记录: {stock_code} {leader_name}")
            return results

        except Exception as e:
            logger.error(f"数据库查询失败: {stock_code} {leader_name} - {e}")
            return []

    def _compare_single_position(self, sql_data_list: List[Dict[str, Any]],
                                  ai_data: Dict[str, Any],
                                  position: str, stock_code: str) -> List[Dict[str, Any]]:
        """比对单个职位的AI数据与SQL数据"""
        comparison_results = []

        if not sql_data_list:
            comparison_results.append(
                self._create_comparison_result({}, ai_data, stock_code,
                                               "正式库无对应代码和领导人的数据，请检查",
                                               position)
            )
            return comparison_results

        ai_zw = self._map_position(position)

        matched = False
        person_exists = False

        for sql_data in sql_data_list:
            sql_xm = str(sql_data.get("XM", "")).strip()
            sql_zwmc = str(sql_data.get("ZWMC", "")).strip()
            sql_zw = str(sql_data.get("ZW", "")).strip()

            if not sql_xm:
                continue

            person_exists = True

            if position and sql_zwmc == position:
                error_messages = self._compare_all_fields(sql_data, ai_data)
                comparison_results.append(
                    self._create_comparison_result(sql_data, ai_data, stock_code,
                                                   error_messages, position)
                )
                matched = True
                break

            if position and ai_zw is not None and sql_zw == ai_zw:
                error_messages = self._compare_all_fields(sql_data, ai_data)
                comparison_results.append(
                    self._create_comparison_result(sql_data, ai_data, stock_code,
                                                   error_messages, position)
                )
                matched = True
                break

        if not matched:
            if not person_exists:
                comparison_results.append(
                    self._create_comparison_result({}, ai_data, stock_code,
                                                   "正式库无对应代码和领导人的数据，请检查",
                                                   position)
                )
            else:
                if position and ai_zw is None:
                    comparison_results.append(
                        self._create_comparison_result({}, ai_data, stock_code,
                                                       "职位无法匹配", position)
                    )
                else:
                    comparison_results.append(
                        self._create_comparison_result({}, ai_data, stock_code,
                                                       "正式库无对应职位", position)
                    )

        return comparison_results

    def _compare_all_fields(self, sql_data: Dict, ai_data: Dict) -> List[str]:
        """全字段比对，返回错误信息列表"""
        error_messages = []
        reverse_mapping = {v: k for k, v in self.field_mapping.items()}

        fields_to_compare = ["XB", "CSRQ", "XL", "GJ"]

        for sql_field in fields_to_compare:
            ai_field = reverse_mapping.get(sql_field, sql_field)

            sql_value = sql_data.get(sql_field)
            ai_value = ai_data.get(ai_field)

            if ai_value is not None and str(ai_value).strip():
                sql_value_str = str(sql_value).strip() if sql_value is not None else ""
                ai_value_str = str(ai_value).strip() if ai_value is not None else ""

                if sql_field == "CSRQ":
                    error_message = self._compare_birthdate(sql_value_str, ai_value_str)
                    if error_message:
                        error_messages.append(error_message)
                else:
                    if sql_value_str != ai_value_str:
                        error_messages.append(f"{ai_field}【AI：{ai_value_str}】")

        return error_messages

    def _compare_birthdate(self, sql_value: str, ai_value: str) -> str:
        """比较出生日期，AI数据精度不超过SQL时不报错"""
        if not ai_value:
            return ""

        try:
            if len(ai_value) > len(sql_value):
                return f"出生日期【AI：{ai_value}】"

            sql_parts = sql_value.split('-')
            ai_parts = ai_value.split('-')

            sql_year = sql_parts[0] if len(sql_parts) > 0 else ''
            sql_month = sql_parts[1] if len(sql_parts) > 1 else ''
            sql_day = sql_parts[2] if len(sql_parts) > 2 else ''

            ai_year = ai_parts[0] if len(ai_parts) > 0 else ''
            ai_month = ai_parts[1] if len(ai_parts) > 1 else ''
            ai_day = ai_parts[2] if len(ai_parts) > 2 else ''

            if ai_year and ai_year != sql_year:
                return f"出生日期【AI：{ai_value}】"
            if ai_month and ai_month != sql_month:
                return f"出生日期【AI：{ai_value}】"
            if ai_day and ai_day != sql_day:
                return f"出生日期【AI：{ai_value}】"

            return ""
        except Exception:
            return f"出生日期【AI：{ai_value}】"

    def _create_comparison_result(self, sql_data: Dict, ai_data: Dict,
                                   stock_code: str, error_description,
                                   position: str = "") -> Dict[str, Any]:
        """创建比对结果对象"""
        if isinstance(error_description, list):
            error_desc = "；".join(error_description) if error_description else ""
        else:
            error_desc = str(error_description) if error_description else ""

        leader_name = ai_data.get("领导人姓名", "") or sql_data.get("XM", "")

        return {
            "ID": str(sql_data.get("ID", "")),
            "股票代码": stock_code,
            "领导人姓名": leader_name,
            "职位描述": position,
            "性别": sql_data.get("XB", ""),
            "出生日期": sql_data.get("CSRQ", ""),
            "学历": sql_data.get("XL", ""),
            "国籍": sql_data.get("GJ", ""),
            "错误描述": error_desc,
            "背景介绍(BJJS)": sql_data.get("BJJS", ""),
            "背景介绍-AI": ai_data.get("背景介绍", ""),
        }

    def _map_position(self, ai_position: str) -> Optional[str]:
        """映射AI职位到数据库职位代码"""
        if not ai_position:
            return None

        standard_zw = self.position_mapping.get(ai_position)
        if standard_zw:
            return standard_zw

        if "委员" in ai_position:
            return "699"

        return None

    def generate_comparison_report(self, all_results: List[Dict[str, Any]],
                                   output_dir: str = None,
                                   session_id: str = "") -> Optional[str]:
        """生成比对结果Excel报告"""
        if not all_results:
            logger.info("数据一致，无需生成比对报告")
            return None

        if output_dir is None:
            output_dir = get_reports_dir()

        os.makedirs(output_dir, exist_ok=True)

        if session_id:
            report_filename = f"比对结果-{session_id}.xlsx"
        else:
            report_filename = "比对结果.xlsx"
        report_path = os.path.join(output_dir, report_filename)

        try:
            df = pd.DataFrame(all_results)

            column_order = [
                "ID", "股票代码", "领导人姓名", "职位描述",
                "性别", "出生日期", "学历", "国籍", "错误描述"
            ]
            existing_cols = [c for c in column_order if c in df.columns]
            df_main = df[existing_cols]

            df_main.to_excel(report_path, index=False, engine='openpyxl')

            from openpyxl import load_workbook
            from openpyxl.cell.text import InlineFont
            from openpyxl.cell.rich_text import TextBlock, CellRichText
            from openpyxl.styles import Alignment
            
            wb = load_workbook(report_path)
            ws = wb.active
            ws.title = "比对结果"

            column_widths = {
                'A': 14, 'B': 10, 'C': 12, 'D': 30,
                'E': 6, 'F': 14, 'G': 8, 'H': 10, 'I': 40
            }
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width

            ws_bjjs = wb.create_sheet("背景介绍")
            
            ws_bjjs.cell(row=1, column=1, value="股票代码")
            ws_bjjs.cell(row=1, column=2, value="姓名")
            ws_bjjs.cell(row=1, column=3, value="背景介绍-正式库")
            ws_bjjs.cell(row=1, column=4, value="背景介绍-AI")
            
            ws_bjjs.column_dimensions['A'].width = 12
            ws_bjjs.column_dimensions['B'].width = 15
            ws_bjjs.column_dimensions['C'].width = 80
            ws_bjjs.column_dimensions['D'].width = 80
            
            seen_names = set()
            row_idx = 2
            
            for result in all_results:
                leader_name = result.get("领导人姓名", "")
                stock_code = result.get("股票代码", "")
                if not leader_name or leader_name in seen_names:
                    continue
                seen_names.add(leader_name)
                
                sql_bjjs = result.get("背景介绍(BJJS)", "")
                ai_bjjs = result.get("背景介绍-AI", "")
                
                preprocessed_sql = preprocess_text(sql_bjjs)
                preprocessed_ai = preprocess_text(ai_bjjs)
                
                sql_parts, ai_parts = compare_texts_char_by_char(preprocessed_sql, preprocessed_ai)
                
                ws_bjjs.cell(row=row_idx, column=1, value=stock_code)
                ws_bjjs.cell(row=row_idx, column=2, value=leader_name)
                
                cell_sql = ws_bjjs.cell(row=row_idx, column=3)
                if sql_parts:
                    rich_text_sql = CellRichText()
                    for text, is_red in sql_parts:
                        if is_red:
                            rich_text_sql.append(TextBlock(InlineFont(color="FF0000"), text))
                        else:
                            rich_text_sql.append(TextBlock(InlineFont(), text))
                    cell_sql.value = rich_text_sql
                else:
                    cell_sql.value = ""
                cell_sql.alignment = Alignment(wrap_text=True, vertical='top')
                
                cell_ai = ws_bjjs.cell(row=row_idx, column=4)
                if ai_parts:
                    rich_text_ai = CellRichText()
                    for text, is_red in ai_parts:
                        if is_red:
                            rich_text_ai.append(TextBlock(InlineFont(color="FF0000"), text))
                        else:
                            rich_text_ai.append(TextBlock(InlineFont(), text))
                    cell_ai.value = rich_text_ai
                else:
                    cell_ai.value = ""
                cell_ai.alignment = Alignment(wrap_text=True, vertical='top')
                
                row_idx += 1

            wb.save(report_path)

            logger.info(f"比对报告已生成: {report_path}")
            print(f"比对报告已生成: {report_path}")
            return report_path

        except Exception as e:
            logger.error(f"生成比对报告失败: {e}")
            return None
